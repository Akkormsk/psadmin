import hmac
import json
import logging
import os
import subprocess
import sys
import time
from datetime import timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_GET, require_POST
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook

from .models import CatalogMatchDecision, CatalogSyncRun, CatalogSupplier, ProcessDefinition, ProductionTrainingExample, ProductionTrainingSession, ProductionTrainingTurn, ProductionType, TenderEstimate, TenderKnowledgeSource, TenderLine, TenderSettings
from .knowledge import export_knowledge_bundle
from .catalog import CatalogSyncError, sync_gifts_catalog
from .services import TenderAIError, _resolve_line_match, analyze_tender_requirements, apply_catalog_candidate, apply_verified_source_quote, build_training_hypothesis, calculate_tender, classify_production_type, detect_tender_document_type, extract_calculation_source, inspect_tender_document, recognize_tender_items, refresh_training_example_embedding


logger = logging.getLogger(__name__)


SUPPORTED_TENDER_DOCUMENTS = {".xlsx", ".xls", ".doc", ".docx", ".pdf"}


def knowledge_sync(request):
    expected = os.getenv("KNOWLEDGE_SYNC_TOKEN", "")
    supplied = request.headers.get("Authorization", "")
    if not expected or not hmac.compare_digest(supplied, f"Bearer {expected}"):
        return HttpResponse(status=403)
    return JsonResponse(export_knowledge_bundle(include_embeddings=True), json_dumps_params={"ensure_ascii": False})


@require_GET
def gifts_import_test(request):
    expected = os.getenv("KNOWLEDGE_SYNC_TOKEN", "")
    supplied = request.headers.get("Authorization", "")
    if not expected or not hmac.compare_digest(supplied, f"Bearer {expected}"):
        return HttpResponse(status=403)
    if request.GET.get("status") == "1":
        run = CatalogSyncRun.objects.filter(supplier__code="gifts").first()
        if run is None:
            return JsonResponse({"status": "not_started"})
        return JsonResponse({"status": run.status, "received": run.received_count, "created": run.created_count, "updated": run.updated_count, "error": run.error})
    full = request.GET.get("full") == "1"
    if full:
        recent_running = CatalogSyncRun.objects.filter(supplier__code="gifts", status="running", started_at__gte=timezone.now() - timedelta(minutes=15)).exists()
        if recent_running:
            return JsonResponse({"status": "already_running"}, status=409)
        subprocess.Popen([sys.executable, "manage.py", "sync_gifts_catalog"], start_new_session=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        return JsonResponse({"status": "started"}, status=202)
    else:
        try:
            limit = int(request.GET.get("limit", "10"))
        except (TypeError, ValueError):
            return JsonResponse({"error": "Параметр limit должен быть числом от 1 до 100."}, status=400)
        if not 1 <= limit <= 100:
            return JsonResponse({"error": "Параметр limit должен быть от 1 до 100."}, status=400)
    started = time.monotonic()
    try:
        run = sync_gifts_catalog(limit=limit)
    except CatalogSyncError as exc:
        return JsonResponse({"error": str(exc)}, status=502)
    products = [
        {"external_id": row["external_id"], "article": row["article"], "name": row["name"]}
        for row in getattr(run, "imported_rows", [])
    ] if not full else []
    return JsonResponse({
        "status": run.status,
        "received": run.received_count,
        "created": run.created_count,
        "updated": run.updated_count,
        "seconds": round(time.monotonic() - started, 2),
        "products": products,
    })

def _document_upload_error(upload):
    if upload is None:
        return "Выберите документ."
    if upload.size > 10 * 1024 * 1024:
        return "Файл больше 10 МБ."
    if Path(upload.name).suffix.lower() not in SUPPORTED_TENDER_DOCUMENTS:
        return "Поддерживаются .xlsx, .xls, .doc, .docx и .pdf."
    return ""


def _estimate_for_user(request, pk):
    estimate = get_object_or_404(TenderEstimate, pk=pk)
    if not request.user.is_superuser and estimate.owner_id != request.user.id:
        raise Http404
    return estimate


def _number(value, default="0"):
    return Decimal(str(value or default).replace(",", "."))


@login_required
@require_POST
def import_preview(request):
    upload = request.FILES.get("file")
    if upload is None or not upload.name.lower().endswith(".xlsx"):
        return JsonResponse({"error": "Выберите файл Excel в формате .xlsx."}, status=400)
    if upload.size > 10 * 1024 * 1024:
        return JsonResponse({"error": "Файл больше 10 МБ."}, status=400)
    try:
        workbook = load_workbook(upload, read_only=True, data_only=True)
        sheet_name = request.POST.get("sheet")
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        if not sheet.max_row or not sheet.max_column:
            sheet.calculate_dimension(force=True)
        max_columns = min(sheet.max_column or 1, 50)
        max_rows = min(sheet.max_row or 1, 500)
        rows = []
        for values in sheet.iter_rows(min_row=1, max_row=max_rows, max_col=max_columns, values_only=True):
            rows.append(["" if value is None else str(value) for value in values])
        while rows and not any(value.strip() for value in rows[-1]):
            rows.pop()
        return JsonResponse({"sheets": workbook.sheetnames, "sheet": sheet.title, "rows": rows, "truncated": (sheet.max_row or 0) > max_rows})
    except Exception:
        return JsonResponse({"error": "Не удалось прочитать файл. Проверьте, что это корректный .xlsx."}, status=400)


@login_required
@require_POST
def document_inspect(request):
    upload = request.FILES.get("file")
    error = _document_upload_error(upload)
    if error:
        return JsonResponse({"error": error}, status=400)
    try:
        return JsonResponse(inspect_tender_document(upload))
    except TenderAIError as exc:
        return JsonResponse({"error": str(exc)}, status=400)
    except Exception:
        return JsonResponse({"error": "Не удалось проверить структуру документа."}, status=400)


@login_required
@require_POST
def ai_import_preview(request):
    upload = request.FILES.get("file")
    error = _document_upload_error(upload)
    if error:
        return JsonResponse({"error": error}, status=400)
    try:
        return JsonResponse(recognize_tender_items(upload))
    except TenderAIError as exc:
        return JsonResponse({"error": str(exc)}, status=400)
    except Exception:
        return JsonResponse({"error": "Не удалось прочитать документ. Проверьте файл и попробуйте ещё раз."}, status=400)


@login_required
@require_POST
def document_preview(request):
    upload = request.FILES.get("file")
    error = _document_upload_error(upload)
    if error:
        return JsonResponse({"error": error}, status=400)
    try:
        raw_lines = json.loads(request.POST.get("lines_json", "[]"))
        current_lines = raw_lines if isinstance(raw_lines, list) else []
    except json.JSONDecodeError:
        current_lines = []
    try:
        requested_role = request.POST.get("document_role", "auto")
        detected_role = detect_tender_document_type(upload)
        if requested_role == "nmck" and detected_role == "technical":
            return JsonResponse({"error": "Этот файл похож на ООЗ или ТЗ. Нажмите «Загрузить ООЗ / ТЗ».", "document_type": detected_role}, status=422)
        if requested_role == "technical" and detected_role == "nmck":
            return JsonResponse({"error": "Этот файл похож на НМЦК. Нажмите «Загрузить НМЦК».", "document_type": detected_role}, status=422)
        role = requested_role if requested_role in {"nmck", "technical"} else detected_role
        result = {"document_type": role, "file_name": upload.name}
        if role in {"nmck", "mixed"}:
            nmck = recognize_tender_items(upload)
            used = set()
            for item in nmck.get("items", []):
                line_index, confidence, reason = _resolve_line_match(None, item.get("name"), item.get("quantity"), current_lines, used)
                item["line_index"] = line_index
                item["match_reason"] = reason
                if line_index is not None:
                    used.add(line_index)
                    item["confidence"] = max(float(item.get("confidence", 0)), confidence)
            result["nmck"] = nmck
        if role in {"technical", "mixed"}:
            result["technical"] = analyze_tender_requirements(upload, current_lines)
        if role == "unknown":
            return JsonResponse({"error": "Не удалось уверенно определить тип документа. Выберите его тип вручную и повторите анализ.", "document_type": role}, status=422)
        return JsonResponse(result)
    except TenderAIError as exc:
        return JsonResponse({"error": str(exc)}, status=400)
    except Exception:
        return JsonResponse({"error": "Не удалось проанализировать документ. Проверьте файл и попробуйте ещё раз."}, status=400)


@login_required
@require_POST
def technical_requirements_preview(request):
    upload = request.FILES.get("file")
    error = _document_upload_error(upload)
    if error:
        return JsonResponse({"error": "Выберите ООЗ или ТЗ." if upload is None else error}, status=400)
    try:
        raw_lines = json.loads(request.POST.get("lines_json", "[]"))
        current_lines = raw_lines if isinstance(raw_lines, list) else []
    except json.JSONDecodeError:
        current_lines = []
    try:
        return JsonResponse(analyze_tender_requirements(upload, current_lines))
    except TenderAIError as exc:
        return JsonResponse({"error": str(exc)}, status=400)
    except Exception:
        return JsonResponse({"error": "Не удалось проанализировать ООЗ/ТЗ. Проверьте файл и попробуйте ещё раз."}, status=400)


@login_required
@require_POST
def production_route_preview(request):
    if not request.user.is_superuser:
        return JsonResponse({"error": "ИИ-расчёт пока доступен только администратору."}, status=403)
    try:
        line = json.loads(request.POST.get("line_json", "{}"))
        if not isinstance(line, dict) or not str(line.get("name", "")).strip() or _number(line.get("quantity")) <= 0:
            raise ValueError
    except (ValueError, TypeError, InvalidOperation, json.JSONDecodeError):
        return JsonResponse({"error": "Сначала заполните позицию и примените требования ТЗ."}, status=400)
    try:
        hypothesis = build_training_hypothesis(line)
        session = ProductionTrainingSession.objects.create(
            created_by=request.user,
            position_name=str(line.get("name", ""))[:500],
            requirements=line.get("requirements") if isinstance(line.get("requirements"), dict) else {},
            current_hypothesis=hypothesis,
        )
        ProductionTrainingTurn.objects.create(session=session, hypothesis=hypothesis)
        hypothesis["session_id"] = session.pk
        return JsonResponse(hypothesis)
    except TenderAIError as exc:
        return JsonResponse({"error": str(exc)}, status=400)
    except Exception:
        logger.exception("Unexpected production hypothesis error")
        return JsonResponse({"error": "Не удалось построить расчёт. Подробная причина записана в журнал приложения."}, status=400)


@login_required
@require_POST
def revise_production_hypothesis(request):
    if not request.user.is_superuser:
        return JsonResponse({"error": "Обучать ассистента может только администратор."}, status=403)
    try:
        payload = json.loads(request.POST.get("payload", "{}"))
        session = ProductionTrainingSession.objects.get(pk=payload.get("session_id"), created_by=request.user, is_confirmed=False)
        line = payload.get("line") if isinstance(payload.get("line"), dict) else {}
        feedback = str(payload.get("feedback", "")).strip()
        if not feedback or len(feedback) > 3000 or not str(line.get("name", "")).strip():
            raise ValueError
    except (ValueError, TypeError, json.JSONDecodeError, ProductionTrainingSession.DoesNotExist):
        return JsonResponse({"error": "Не удалось продолжить диалог. Обновите гипотезу и повторите."}, status=400)
    try:
        hypothesis = build_training_hypothesis(line, current=session.current_hypothesis, feedback=feedback)
        hypothesis["session_id"] = session.pk
        session.position_name = str(line.get("name", ""))[:500]
        session.requirements = line.get("requirements") if isinstance(line.get("requirements"), dict) else {}
        session.current_hypothesis = hypothesis
        session.save(update_fields=["position_name", "requirements", "current_hypothesis", "updated_at"])
        ProductionTrainingTurn.objects.create(
            session=session,
            feedback=feedback,
            understood_changes=hypothesis.get("understood_changes", []),
            hypothesis=hypothesis,
        )
        return JsonResponse(hypothesis)
    except TenderAIError as exc:
        return JsonResponse({"error": str(exc)}, status=400)
    except Exception:
        return JsonResponse({"error": "Не удалось учесть исправление. Попробуйте ещё раз."}, status=400)


@login_required
@require_POST
def select_catalog_product(request):
    if not request.user.is_superuser:
        return JsonResponse({"error": "Выбирать товары для обучения может только администратор."}, status=403)
    try:
        payload = json.loads(request.POST.get("payload", "{}"))
        session = ProductionTrainingSession.objects.get(pk=payload.get("session_id"), created_by=request.user, is_confirmed=False)
        line = payload.get("line") if isinstance(payload.get("line"), dict) else {}
        product_id = str(payload.get("product_id", "")).strip()[:100]
        if not product_id:
            raise ValueError
        if not str(line.get("name", "")).strip():
            raise ValueError
    except (ValueError, TypeError, json.JSONDecodeError, ProductionTrainingSession.DoesNotExist):
        return JsonResponse({"error": "Не удалось выбрать товар поставщика. Обновите гипотезу."}, status=400)
    try:
        hypothesis = apply_catalog_candidate(session.current_hypothesis, line, product_id)
        hypothesis["session_id"] = session.pk
        session.current_hypothesis = hypothesis
        session.save(update_fields=["current_hypothesis", "updated_at"])
        selection = hypothesis.get("catalog_selection", {})
        CatalogMatchDecision.objects.create(
            session=session,
            product=None,
            supplier_code=str(selection.get("supplier_code") or "oasis")[:50],
            product_external_id=str(selection.get("external_id") or selection.get("id") or "")[:100],
            product_article=str(selection.get("article") or "")[:120],
            product_snapshot={
                key: selection.get(key) for key in (
                    "external_id", "article", "name", "price", "stock", "url", "category", "matches",
                    "supplier_code", "supplier_name", "supplier_site",
                )
            },
            decision="selected",
            reason_codes=[
                "backend_exact_match" if selection.get("fit") == "exact" else "admin_override_mismatch",
                "selected_by_admin",
            ],
            requirement_signature={
                "name": str(line.get("name", ""))[:500],
                "quantity": str(line.get("quantity", ""))[:50],
                "requirements": line.get("requirements") if isinstance(line.get("requirements"), dict) else {},
            },
            created_by=request.user,
        )
        supplier_name = str(selection.get("supplier_name") or selection.get("supplier_code") or "поставщика")
        change = f"Выбран товар поставщика {supplier_name}: {selection.get('name', 'товар')} · арт. {selection.get('article', '')}".strip()
        ProductionTrainingTurn.objects.create(session=session, feedback=change, understood_changes=[change], hypothesis=hypothesis)
        return JsonResponse(hypothesis)
    except TenderAIError as exc:
        return JsonResponse({"error": str(exc)}, status=400)
    except Exception:
        return JsonResponse({"error": "Не удалось применить товар поставщика к расчёту."}, status=400)


@login_required
@require_POST
def add_calculation_source(request):
    if not request.user.is_superuser:
        return JsonResponse({"error": "Добавлять источники расчёта может только администратор."}, status=403)
    try:
        payload = json.loads(request.POST.get("payload", "{}"))
        session = ProductionTrainingSession.objects.get(pk=payload.get("session_id"), created_by=request.user, is_confirmed=False)
        line = payload.get("line") if isinstance(payload.get("line"), dict) else {}
        hypothesis = session.current_hypothesis if isinstance(session.current_hypothesis, dict) else {}
        costs = hypothesis.get("costs") if isinstance(hypothesis.get("costs"), list) else []
        if not str(line.get("name", "")).strip():
            raise ValueError
        raw_sources = payload.get("sources") if isinstance(payload.get("sources"), list) else [payload]
        if not raw_sources or len(raw_sources) > 6:
            raise ValueError
    except (ValueError, TypeError, json.JSONDecodeError, ProductionTrainingSession.DoesNotExist, TenderKnowledgeSource.DoesNotExist):
        return JsonResponse({"error": "Не удалось определить статью расчёта или источник."}, status=400)

    try:
        prepared_sources = []
        batch_mode = isinstance(payload.get("sources"), list)
        administrator_feedback = str(payload.get("feedback", "")).strip()[:6000]
        for source_index, raw_source in enumerate(raw_sources):
            if not isinstance(raw_source, dict):
                raise ValueError
            raw_cost_index = raw_source.get("cost_index")
            cost_index = int(raw_cost_index) if raw_cost_index not in (None, "") else None
            if cost_index is not None and (cost_index < 0 or cost_index >= len(costs)):
                raise ValueError
            upload = request.FILES.get(f"file_{source_index}" if batch_mode else "file")
            if upload is not None and upload.size > 10 * 1024 * 1024:
                return JsonResponse({"error": f"Файл источника № {source_index + 1} больше 10 МБ."}, status=400)
            existing = None
            target_name = costs[cost_index].get("name", "") if cost_index is not None else ""
            if raw_source.get("source_id"):
                existing = TenderKnowledgeSource.objects.get(pk=raw_source["source_id"], is_active=True)
            if existing:
                if existing.url:
                    extracted = extract_calculation_source(
                        source_url=existing.url,
                        selection_context={"line": line, "feedback": administrator_feedback, "target_name": target_name},
                    )
                    refreshed = extracted.get("structured_data") if isinstance(extracted.get("structured_data"), dict) else {}
                    existing.content_summary = extracted["content"]
                    existing.structured_data = {
                        **(existing.structured_data if isinstance(existing.structured_data, dict) else {}),
                        **refreshed,
                    }
                    existing.save(update_fields=["content_summary", "structured_data", "updated_at"])
                else:
                    extracted = {
                        "content": existing.content_summary,
                        "source_type": existing.source_type,
                        "url": existing.url,
                        "structured_data": existing.structured_data if isinstance(existing.structured_data, dict) else {},
                    }
                source = existing
            else:
                extracted = extract_calculation_source(
                    source_text=str(raw_source.get("source_text", "")),
                    source_url=str(raw_source.get("source_url", "")),
                    upload=upload,
                    selection_context={"line": line, "feedback": administrator_feedback, "target_name": target_name},
                )
                title = str(raw_source.get("title", "")).strip() or str(raw_source.get("supplier_name", "")).strip() or target_name or f"Источник для {line.get('name')}"
                extracted_structured = extracted.get("structured_data") if isinstance(extracted.get("structured_data"), dict) else {}
                source = TenderKnowledgeSource.objects.create(
                    title=title[:300],
                    supplier_name=str(raw_source.get("supplier_name", "")).strip()[:200],
                    source_type=extracted["source_type"],
                    url=extracted["url"],
                    content_summary=extracted["content"],
                    structured_data={
                        "scope": "cost" if cost_index is not None else "position",
                        "position_name": str(line.get("name", ""))[:300],
                        "cost_name": target_name,
                        **extracted_structured,
                    },
                    created_by=request.user,
                    is_active=False,
                )
            prepared_sources.append({
                "source": source,
                "extracted": extracted,
                "cost_index": cost_index,
                "target": costs[cost_index] if cost_index is not None else None,
            })

        feedback_parts = [administrator_feedback] if administrator_feedback else []
        for source_index, prepared in enumerate(prepared_sources, start=1):
            source, target, extracted = prepared["source"], prepared["target"], prepared["extracted"]
            if target is not None:
                instruction = (
                    f"Для статьи «{target.get('name', 'расход')}» добавлен проверяемый источник «{source}». "
                    "Пересчитай эту статью по данным источника, не копируй итог из похожего заказа."
                )
            else:
                instruction = (
                    f"К позиции добавлен проверяемый источник «{source}». Сам определи, к какому процессу, "
                    "материалу или статье цены он относится. Используй его только в подходящем маршруте; "
                    "наличие источника само по себе не подтверждает маршрут и не делает его оптимальным."
                )
            feedback_parts.append(f"ИСТОЧНИК № {source_index}. {instruction}\nДАННЫЕ ИСТОЧНИКА:\n{extracted['content'][:12000]}")
        feedback_parts.append("Сохрани универсальные процессы, подробные формулы, все промежуточные действия и способы адаптации к текущему тиражу. Не смешивай предложения разных поставщиков в одну цену.")
        feedback = "\n\n".join(feedback_parts)
        updated = build_training_hypothesis(line, current=hypothesis, feedback=feedback)
        for prepared in prepared_sources:
            structured = prepared["extracted"].get("structured_data") if isinstance(prepared["extracted"].get("structured_data"), dict) else {}
            updated = apply_verified_source_quote(
                updated,
                line,
                structured.get("price_quote"),
                str(prepared["source"]),
                prepared["source"].url,
                prepared["target"],
            )
        updated["session_id"] = session.pk
        attached_sources = hypothesis.get("sources") if isinstance(hypothesis.get("sources"), list) else []
        new_source_ids = {prepared["source"].pk for prepared in prepared_sources}
        source_cards = [{
            "id": prepared["source"].pk,
            "title": prepared["source"].title,
            "supplier_name": prepared["source"].supplier_name,
            "source_type": prepared["source"].source_type,
            "url": prepared["source"].url,
            "scope": "cost" if prepared["target"] is not None else "position",
            "cost_name": prepared["target"].get("name", "") if prepared["target"] is not None else "",
            "is_pending": not prepared["source"].is_active,
        } for prepared in prepared_sources]
        updated["sources"] = [value for value in attached_sources if value.get("id") not in new_source_ids] + source_cards
        updated_costs = updated.get("costs") if isinstance(updated.get("costs"), list) else []
        for prepared in prepared_sources:
            source, target, cost_index = prepared["source"], prepared["target"], prepared["cost_index"]
            if target is None:
                continue
            matching = next((item for item in updated_costs if str(item.get("name", "")).casefold() == str(target.get("name", "")).casefold()), None)
            if matching is None and cost_index < len(updated_costs):
                matching = updated_costs[cost_index]
            if matching is not None:
                structured = prepared["extracted"].get("structured_data") if isinstance(prepared["extracted"].get("structured_data"), dict) else {}
                matching.update({
                    "source": str(source), "source_id": source.pk,
                    "source_type": "supplier" if structured.get("price_quote") else source.source_type,
                    "source_url": source.url, "source_date": source.updated_at.date().isoformat(),
                })
        session.current_hypothesis = updated
        session.save(update_fields=["current_hypothesis", "updated_at"])
        ProductionTrainingTurn.objects.create(
            session=session, feedback=feedback, understood_changes=updated.get("understood_changes", []), hypothesis=updated,
        )
        return JsonResponse(updated)
    except TenderAIError as exc:
        return JsonResponse({"error": str(exc)}, status=400)
    except Exception:
        return JsonResponse({"error": "Не удалось прочитать источник или пересчитать статью."}, status=400)


@login_required
@require_POST
def confirm_production_type(request):
    if not request.user.is_superuser:
        return JsonResponse({"error": "Добавлять учебные примеры может только администратор."}, status=403)
    try:
        payload = json.loads(request.POST.get("payload", "{}"))
        line = payload.get("line") if isinstance(payload.get("line"), dict) else {}
        session_id = payload.get("session_id")
        if session_id:
            session = ProductionTrainingSession.objects.get(pk=session_id, created_by=request.user, is_confirmed=False)
            hypothesis = session.current_hypothesis if isinstance(session.current_hypothesis, dict) else {}
            production_type = ProductionType.objects.get(code=hypothesis.get("product_type"), is_active=True)
            name = session.position_name.strip()
            if not name or not hypothesis.get("route"):
                raise ValueError
            learning_warnings = [str(value) for value in hypothesis.get("learning_warnings", []) if str(value).strip()]
            if learning_warnings:
                return JsonResponse({
                    "error": "Пока нельзя сохранить в обучение: " + " ".join(learning_warnings[:3]),
                    "learning_warnings": learning_warnings,
                }, status=400)
            route = hypothesis["route"]
            attached_source_ids = [
                value.get("id") for value in hypothesis.get("sources", [])
                if isinstance(value, dict) and value.get("id")
            ]
            example = ProductionTrainingExample.objects.create(
                production_type=production_type,
                position_name=name[:500],
                requirements=session.requirements,
                features=hypothesis.get("facts", [])[:10],
                routes=[{
                    "name": str(route.get("name", ""))[:200],
                    "reason": str(route.get("reason", ""))[:700],
                    "steps": route.get("steps", [])[:6],
                    "processes": route.get("processes", [])[:6],
                    "costs": hypothesis.get("costs", [])[:12],
                    "totals": hypothesis.get("totals", {}),
                    "psodin_calculation": hypothesis.get("psodin_calculation", {}),
                    "catalog_intent": hypothesis.get("catalog_intent", {}),
                    "catalog_selection": hypothesis.get("catalog_selection", {}),
                }],
                note=str(payload.get("note", ""))[:500],
                created_by=request.user,
            )
            ProductionTrainingExample.objects.filter(
                production_type=production_type,
                position_name__iexact=name,
                is_active=True,
            ).exclude(pk=example.pk).update(is_active=False, superseded_by=example)
            refresh_training_example_embedding(example)
            if attached_source_ids:
                TenderKnowledgeSource.objects.filter(
                    pk__in=attached_source_ids, created_by=request.user, is_active=False,
                ).update(is_active=True)
            session.is_confirmed = True
            session.confirmed_example = example
            session.save(update_fields=["is_confirmed", "confirmed_example", "updated_at"])
            session.catalog_decisions.update(is_confirmed=True)
            return JsonResponse({"message": f"Расчёт подтверждён и сохранён как учебный пример: {production_type.name}.", "example_id": example.pk})
        production_type = ProductionType.objects.get(code=payload.get("production_type"), is_active=True)
        name = str(line.get("name", "")).strip()
        if not name:
            raise ValueError
        features = payload.get("features") if isinstance(payload.get("features"), list) else []
        raw_routes = payload.get("routes") if isinstance(payload.get("routes"), list) else []
    except (ValueError, TypeError, json.JSONDecodeError, ProductionType.DoesNotExist, ProductionTrainingSession.DoesNotExist):
        return JsonResponse({"error": "Не удалось сохранить учебный пример."}, status=400)
    routes = []
    for route_index, raw_route in enumerate(raw_routes[:5]):
        if not isinstance(raw_route, dict):
            continue
        processes = []
        for raw_process in raw_route.get("processes", [])[:12]:
            if not isinstance(raw_process, dict):
                continue
            role = raw_process.get("role") if raw_process.get("role") in {"supply", "production", "completion"} else "production"
            process_name = str(raw_process.get("name", "")).strip()[:200]
            if not process_name:
                continue
            ProcessDefinition.objects.get_or_create(name=process_name, role=role)
            processes.append({"role": role, "name": process_name, "reason": str(raw_process.get("reason", ""))[:300]})
        if processes:
            routes.append({"name": str(raw_route.get("name", "")).strip()[:120] or f"Маршрут {route_index + 1}", "reason": str(raw_route.get("reason", ""))[:300], "processes": processes})
    if not routes:
        return JsonResponse({"error": "Добавьте хотя бы один процесс в маршрут."}, status=400)
    example = ProductionTrainingExample.objects.create(
        production_type=production_type,
        position_name=name[:500],
        requirements=line.get("requirements") if isinstance(line.get("requirements"), dict) else {},
        features=[str(value)[:300] for value in features[:10]],
        routes=routes,
        note=str(payload.get("note", ""))[:500],
        created_by=request.user,
    )
    refresh_training_example_embedding(example)
    ProductionTrainingExample.objects.filter(
        production_type=production_type,
        position_name__iexact=name,
        is_active=True,
    ).exclude(pk=example.pk).update(is_active=False, superseded_by=example)
    return JsonResponse({"message": f"Пример сохранён: {production_type.name}.", "example_id": example.pk})


@login_required
@require_POST
def calculator_knowledge_proposal(request):
    return JsonResponse({"error": "Добавьте постоянный расходник через калькулятор или админку."}, status=410)


@login_required
def home(request, pk=None):
    estimate = _estimate_for_user(request, pk) if pk else None
    settings = TenderSettings.objects.get_or_create(pk=1)[0]
    posted_lines = None
    posted_analysis = None
    form_state = {
        "tender_number": estimate.tender_number if estimate else "",
        "name": estimate.name if estimate else "",
        "reduction_percent": estimate.reduction_percent if estimate else Decimal("30.00"),
        "russia_delivery": estimate.russia_delivery if estimate else Decimal("0.00"),
        "owner_id": estimate.owner_id if estimate else request.user.id,
    }
    if request.method == "POST":
        form_state = {
            "tender_number": request.POST.get("tender_number", ""),
            "name": request.POST.get("name", ""),
            "reduction_percent": request.POST.get("reduction_percent", "30"),
            "russia_delivery": request.POST.get("russia_delivery", "0"),
            "owner_id": request.POST.get("owner_id") or request.user.id,
        }
        try:
            tender_number = request.POST.get("tender_number", "").strip()
            name = request.POST.get("name", "").strip()
            reduction_percent = _number(request.POST.get("reduction_percent"), "30")
            russia_delivery = _number(request.POST.get("russia_delivery"))
            raw_lines = json.loads(request.POST.get("lines_json", "[]"))
            raw_analysis = json.loads(request.POST.get("document_analysis_json", "{}"))
            posted_lines = raw_lines if isinstance(raw_lines, list) else []
            posted_analysis = raw_analysis if isinstance(raw_analysis, dict) else {}
            if not tender_number or not name or not Decimal("0") <= reduction_percent <= Decimal("100") or russia_delivery < 0:
                raise ValueError
            lines = []
            calculation_complete = True
            for row in raw_lines:
                line = {
                    "name": str(row.get("name", "")).strip(),
                    "quantity": _number(row.get("quantity")),
                    "nmck_unit": _number(row.get("nmck_unit")),
                    "material_unit": _number(row.get("material_unit")),
                    "application_unit": _number(row.get("application_unit")),
                    "logistics_unit": _number(row.get("logistics_unit")),
                    "product_url": str(row.get("product_url", "")).strip(),
                    "comment": str(row.get("comment", "")).strip(),
                    "requirements": row.get("requirements") if isinstance(row.get("requirements"), dict) else {},
                }
                has_expense = any(line[key] > 0 for key in ("material_unit", "application_unit", "logistics_unit"))
                if min(line["quantity"], line["nmck_unit"], line["material_unit"], line["application_unit"], line["logistics_unit"]) < 0:
                    raise ValueError
                if not line["name"] or line["quantity"] <= 0 or line["nmck_unit"] <= 0 or not has_expense:
                    calculation_complete = False
                lines.append(line)
            if not lines:
                lines.append({"name": "", "quantity": Decimal("0"), "nmck_unit": Decimal("0"), "material_unit": Decimal("0"), "application_unit": Decimal("0"), "logistics_unit": Decimal("0"), "product_url": "", "comment": "", "requirements": {}})
                calculation_complete = False
        except (ValueError, TypeError, InvalidOperation, json.JSONDecodeError):
            messages.error(request, "Проверьте реквизиты тендера и товарные позиции.")
        else:
            calculated, summary = calculate_tender(lines, reduction_percent, russia_delivery, settings.vat_rate)
            estimate = estimate or TenderEstimate(owner=request.user)
            if request.user.is_superuser and request.POST.get("owner_id"):
                estimate.owner = get_object_or_404(get_user_model(), pk=request.POST["owner_id"])
            estimate.tender_number = tender_number[:100]
            estimate.name = name[:300]
            estimate.reduction_percent = reduction_percent
            estimate.russia_delivery = russia_delivery
            estimate.vat_rate_snapshot = settings.vat_rate
            estimate.summary_snapshot = {key: str(value) for key, value in summary.items()}
            estimate.summary_snapshot["is_incomplete"] = not calculation_complete
            estimate.document_analysis = posted_analysis or {}
            estimate.save()
            estimate.lines.all().delete()
            TenderLine.objects.bulk_create([TenderLine(estimate=estimate, sort_order=index, **line) for index, line in enumerate(lines)])
            messages.success(request, "Черновик просчёта сохранён." if not calculation_complete else "Просчёт тендера сохранён.")
            return redirect("tender_estimate", pk=estimate.pk)

    initial_lines = []
    if posted_lines is not None:
        initial_lines = posted_lines
    elif estimate:
        initial_lines = [{"name": line.name, "quantity": str(line.quantity), "nmck_unit": str(line.nmck_unit), "material_unit": str(line.material_unit), "application_unit": str(line.application_unit), "logistics_unit": str(line.logistics_unit), "product_url": line.product_url, "comment": line.comment, "requirements": line.requirements} for line in estimate.lines.all()]
    initial_analysis = posted_analysis if posted_analysis is not None else (estimate.document_analysis if estimate else {})
    estimates = TenderEstimate.objects.all() if request.user.is_superuser else TenderEstimate.objects.filter(owner=request.user)
    users = get_user_model().objects.filter(is_active=True).order_by("last_name", "first_name", "username") if request.user.is_superuser else None
    knowledge_sources = []
    if request.user.is_superuser:
        knowledge_sources = list(TenderKnowledgeSource.objects.filter(is_active=True).values("id", "title", "supplier_name", "source_type", "url")[:100])
    return render(request, "tenders/home.html", {"estimate": estimate, "form_state": form_state, "estimates": estimates.select_related("owner", "owner__profile")[:30], "initial_lines_json": json.dumps(initial_lines, ensure_ascii=False), "initial_analysis_json": json.dumps(initial_analysis, ensure_ascii=False), "knowledge_sources_json": json.dumps(knowledge_sources, ensure_ascii=False), "vat_rate": settings.vat_rate, "users": users})


@login_required
@require_POST
def delete_estimate(request, pk):
    estimate = _estimate_for_user(request, pk)
    estimate.delete()
    messages.success(request, "Просчёт тендера удалён.")
    return redirect("tender_home")
